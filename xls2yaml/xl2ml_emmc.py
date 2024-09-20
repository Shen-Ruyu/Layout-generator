import argparse
import openpyxl
import yaml

from .xl2ml_common import yamlData
from .xl2ml_common import Size
from .xl2ml_common import str2hex

# eMMC

SHEET="eMMC layout"
NANE_COLUMN = "Partitions"
PARTITION_COLUMN = "Partition number"
START_LBA_COLUMN = "Start LBA"
PARTITION_FS_COLUMN = "Partition filesystem"
START_ADDRESS_COLUMN = "Start address (0x)"
END_ADDRESS_COLUMN = "End address  + 1 (0x)"
SIZE_COLUMN = "Partiton size (sectors)"
LBID_COLUMN = "LB-ID\n(hex)"
UPDATE_COLUMN = "Update Part (aka LB)"

PASS=[
        "freespace",
        "free space",
        # "Kernel Linux (flash)",
        # "DTB (flash)",
        # "rootfs(flash)",
      ]

class Block:
    """
        Store block values in layout map
    """
    name=""
    size=""
    start_address=""

    def __init__(self,name,size,start_address,end_address,partition,lba,fs,lb_id,update,bank) :
        self.block={
                "name":name,
                "size":size,
                "start_address":start_address,
                "end_address":end_address,
                "partition":partition,
                "lba":lba,
                "fs":fs,
                "lb_id":lb_id,
                "update":update,
                "bank":bank,
        }
        
        print(f"Add:\t\t{self.block}")

class xlData_emmc:
    name=""
    version=""
    project=""
    start_address=""
    size=""

    def __init__(self,name,version,project,start_address,size) :
        self.xl_datas={
            "name": name,
            "version": version,
            "project": project,
            "start_address": start_address,
            "size": size,
            "blocks":[]
            }
        self.cols={
            "NameCol":0,
            "PartitionCol":0,
            "PartitionFSCol":0,
            "StartAddrCol":0,
            "EndCol":0,
            "LBIDCol":0,
            "UpdateCol":0,
            "StartLBACol":0,
        }
    
    def Readxl(self,xlFile,SheetName):
        """
        Parse layout and store the result
        """
        BookData=openpyxl.load_workbook(xlFile,data_only=True)
        self.Sheet=BookData[SheetName]
        print(self.Sheet)
        
        for c in range(1,self.Sheet.max_column+1):
            cell_value=self.Sheet.cell(row=1,column=c).value
            if cell_value == NANE_COLUMN:
                self.cols["NameCol"]=c
            elif cell_value == PARTITION_COLUMN:
                self.cols["PartitionCol"]=c
            elif cell_value == PARTITION_FS_COLUMN:
                self.cols["PartitionFSCol"]=c
            elif cell_value == START_ADDRESS_COLUMN:
                self.cols["StartAddrCol"]=c
            elif cell_value == END_ADDRESS_COLUMN:
                self.cols["EndCol"]=c
            elif cell_value == LBID_COLUMN:
                self.cols["LBIDCol"]=c
            elif cell_value == UPDATE_COLUMN:
                self.cols["UpdateCol"]=c
            elif cell_value == START_LBA_COLUMN:
                self.cols["StartLBACol"]=c

    
        for col in self.cols.keys():
            if self.cols[col]:
                print(f"{col}: {self.cols[col]}")
            else:
                print(f"Can't Find {col}")

        bank=""
        for r in range(2,self.Sheet.max_row+1):
            cell_value=self.Sheet.cell(row=r,column=self.cols["NameCol"]).value
            if cell_value in PASS :
                pass            
            elif type(cell_value)==str:
                blocks_name=cell_value.replace(" ","_").replace("-","_").replace("(","_").replace(")","").replace("+","_").replace("__","_").replace("__","_")
                
                blocks_end_address=self.Sheet.cell(row=r,column=self.cols["EndCol"]).value
                blocks_start_address=self.Sheet.cell(row=r,column=self.cols["StartAddrCol"]).value
                blocks_partition=self.Sheet.cell(row=r,column=self.cols["PartitionCol"]).value
                blocks_lba=self.Sheet.cell(row=r,column=self.cols["StartLBACol"]).value
                blocks_fs=self.Sheet.cell(row=r,column=self.cols["PartitionFSCol"]).value
                blocks_lb_id=self.Sheet.cell(row=r,column=self.cols["LBIDCol"]).value
                blocks_update=self.Sheet.cell(row=r,column=self.cols["UpdateCol"]).value
                
                if blocks_end_address and blocks_start_address !=  None:
                    blocks_start_address=str2hex.FormatAddr(blocks_start_address)
                    blocks_end_address=str2hex.FormatAddr(blocks_end_address)
                    blocks_size=int(blocks_end_address,16)-int(blocks_start_address,16)+1
                    blocks_size=Size.B2KB2MB(int(blocks_size))

                    blocks_start_address=str2hex.FormatAddr(self.Sheet.cell(row=r,column=self.cols["StartAddrCol"]).value)

                    self.xl_datas["blocks"].append(Block(blocks_name,blocks_size,blocks_start_address,blocks_end_address,blocks_partition,blocks_lba,blocks_fs,blocks_lb_id,blocks_update,bank).block)
                else:
                    bank=cell_value.replace(" ","_").replace("(","_").replace(")","").replace("/","_").replace("__","_").replace("__","_")
                    # print(f"{bank}{'='*150}")
        print("<<<Done")
           
