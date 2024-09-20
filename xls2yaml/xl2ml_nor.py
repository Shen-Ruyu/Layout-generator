import argparse
import openpyxl
import yaml

from .xl2ml_common import yamlData
from .xl2ml_common import Size
from .xl2ml_common import str2hex

# NOR
# """
SHEET="Partitioning"
NANE_COLUMN = "Component"
SUB_COLUMN="Subsection"
START_ADDRESS_COLUMN = "Start address (0x)"
END_ADDRESS_COLUMN = "End address (0x)"
SIZE_COLUMN = "Partition size (B)"
LBID_COLUMN = "LB-ID\n(Hex)"
UPDATE_COLUMN = "Update parts (LBs on PDX)"
PASS=[
    #   "Conti/OEM owned constant EOL data written only once at manufaturing of the device",
    #   "this is the active part of the control area used during operation. The main use case is software update",
    #   "this is the backup  part of the control area used before the active part gets updated.",
    #   "Continental owned constant EOL data written only once at manufaturing of the device",
    #   "OEM owned constant EOL data written only once at manufaturing of the device",
      "Free space",
      "reserved",
      " Reserved",
      "Reserved",
      "Free space at end of NOR"
      ]
# """

class Block:
    """
        Store block values in layout map
    """
    name=""
    size=""
    start_address=""

    def __init__(self,name,size,start_address,end_address,lb_id,update,bank) :
        self.block={
                "name":name,
                "size":size,
                "start_address":start_address,
                "end_address":end_address,
                "lb_id":lb_id,
                "update":update,
                "bank":bank
        }
        
        # print(f"Add:\t\t{self.block}")

class xlData_nor:
    name=""
    version=""
    project=""
    start_address=""
    size=""

    def __init__(self,name,version,project,start_address,size) :
        self.xl_datas={
            "name": name,
            "version": version,
            "project": project,
            "start_address": start_address,
            "size": size,
            "blocks":[]
            }
        self.cols={
            "NameCol":0,
            "SubCol":0,
            "StartAddrCol":0,
            "EndCol":0,
            "LBIDCol":0,
            "UpdateCol":0,
        }
    
    def Readxl(self,xlFile,SheetName):
        """
        Parse layout and store the result
        """
        BookData=openpyxl.load_workbook(xlFile,data_only=True)
        self.Sheet=BookData[SheetName]
        print(self.Sheet)
        
        for c in range(1,self.Sheet.max_column+1):
            cell_value=self.Sheet.cell(row=1,column=c).value
            if cell_value == NANE_COLUMN:
                self.cols["NameCol"]=c
            elif cell_value == SUB_COLUMN:
                self.cols["SubCol"]=c
            elif cell_value == START_ADDRESS_COLUMN:
                self.cols["StartAddrCol"]=c
            elif cell_value == END_ADDRESS_COLUMN:
                self.cols["EndCol"]=c
            elif cell_value == LBID_COLUMN:
                self.cols["LBIDCol"]=c
            elif cell_value == UPDATE_COLUMN:
                self.cols["UpdateCol"]=c
      
        for col in self.cols.keys():
            if self.cols[col]:
                print(f"{col}: {self.cols[col]}")
            else:
                print(f"Can't Find {col}")

        bank=""
        pre_name=""
        for r in range(2,self.Sheet.max_row+1):
            cell_value=self.Sheet.cell(row=r,column=self.cols["NameCol"]).value
            if cell_value in PASS :
                pass            
            elif cell_value=="OVERVIEW":
                print("<<<Done")
                break
            else:
                if type(cell_value)==str and len(cell_value)<50:
                    blocks_name=cell_value.replace("(U-boot)","")
                    blocks_name=blocks_name.replace(" ","").replace("-","_").replace("(","_").replace(")","").replace("+","_").replace("__","_").replace("__","_")
                    pre_name=blocks_name
                elif self.Sheet.cell(row=r,column=self.cols["SubCol"]).value:
                    blocks_sub=self.Sheet.cell(row=r,column=self.cols["SubCol"]).value
                    blocks_sub="_"+blocks_sub.replace(" ","")
                    blocks_name=pre_name+blocks_sub
                    blocks_name=blocks_name.replace(" ","").replace("-","_").replace("(","_").replace(")","").replace("+","_").replace("__","_").replace("__","_")
                blocks_end_address=self.Sheet.cell(row=r,column=self.cols["EndCol"]).value
                blocks_start_address=self.Sheet.cell(row=r,column=self.cols["StartAddrCol"]).value
                blocks_lb_id=self.Sheet.cell(row=r,column=self.cols["LBIDCol"]).value
                blocks_update=self.Sheet.cell(row=r,column=self.cols["UpdateCol"]).value
                
                if blocks_end_address !=  None and blocks_start_address !=  None:
                    blocks_start_address=str2hex.FormatAddr(blocks_start_address)
                    blocks_end_address=str2hex.FormatAddr(blocks_end_address)
                    blocks_size=int(blocks_end_address,16)-int(blocks_start_address,16)+1
                    blocks_size=Size.B2KB2MB(int(blocks_size))

                    blocks_start_address=str2hex.FormatAddr(self.Sheet.cell(row=r,column=self.cols["StartAddrCol"]).value)

                    self.xl_datas["blocks"].append(Block(blocks_name,blocks_size,blocks_start_address,blocks_end_address,blocks_lb_id,blocks_update,bank).block)
                else:
                    bank=blocks_name.replace(" ","")
          
